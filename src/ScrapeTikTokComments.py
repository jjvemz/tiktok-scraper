#!/usr/bin/env python3
import sys
from csv import reader
from os import system, remove, path
from datetime import datetime as d
from pyperclip import paste, PyperclipException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import requests
import re
import time

cur_dir = path.dirname(path.abspath(__file__))
csv_path = path.join(cur_dir, "..", "Comments.csv")
system("")

try:
    csv = paste()
except PyperclipException:
    print("\x1b[31m[*]\x1b[0m Could not find copy/paste mechanism on this system. Please paste the csv below and end the input with an empty line:")
    aux = ''
    csv = '\n'.join(iter(input, aux))

try:
    open(csv_path, "w", encoding="utf-8").write(csv.replace("\r","\n").replace("\n\n","\n"))
except Exception as e:
    print(e)
    sys.exit(1)

wb = Workbook()
ws = wb.active

# Función para obtener seguidores de un usuario (mejorada)
def get_followers(username):
    if username.startswith('@'):
        username = username[1:]
    
    url = f"https://www.tiktok.com/@{username}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Referer': 'https://www.tiktok.com/',
        'Cache-Control': 'no-cache'
    }
    
    try:
        print(f"Obteniendo seguidores para: @{username}")
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            # Buscar el patrón de followerCount
            match = re.search(r'"followerCount":(\d+)', r.text)
            if match:
                followers = int(match.group(1))
                print(f"  ✓ {followers:,} seguidores")
                return followers
            
            # Patrón alternativo
            alt_match = re.search(r'follower[^"]*":(\d+)', r.text, re.IGNORECASE)
            if alt_match:
                followers = int(alt_match.group(1))
                print(f"  ✓ {followers:,} seguidores (patrón alternativo)")
                return followers
                
            print(f"  ✗ No se encontraron datos de seguidores en la respuesta")
            return "N/A"
        else:
            print(f"  ✗ Error HTTP {r.status_code}")
            return "N/A"
    except Exception as e:
        print(f"  ✗ Error: {str(e)}")
        return "N/A"

def format_date_for_filename(date_str):
    """Convierte una fecha del CSV al formato para el nombre del archivo"""
    try:
        # Si la fecha está en formato DD-MM-YYYY
        if '-' in date_str and len(date_str.split('-')) == 3:
            parts = date_str.split('-')
            return f"{parts[0]}-{parts[1]}-{parts[2]}"
        else:
            # Si es otro formato, usamos tal como está
            return date_str.replace('/', '-').replace(' ', '_')
    except:
        return "unknown_date"

# Leer y procesar CSV
print("Procesando archivo CSV...")

# Variables para almacenar los datos
metadata = {}  # Para las primeras 14 filas
comments_data = []  # Para los datos de comentarios
publish_time = ""

with open(csv_path, 'r', encoding="utf-8") as f:
    csv_reader = reader(f)
    
    metadata_keys = [
        "Now", "Post URL", "Publisher Nickname", "Publisher @", "Publisher URL",
        "Publish Time", "Post Likes", "Post Shares", "Description",
        "Number of 1st level comments", "Number of 2nd level comments",
        "Total Comments (actual, in this list, rendered in the comment section; needs all comments to be loaded!)",
        "Total Comments (which TikTok tells you; it's too high most of the time when dealing with many comments OR way too low because TikTok limits the number of comments to prevent scraping)",
        "Difference"
    ]
    
    for row_num, row in enumerate(csv_reader):
        if len(row) < 2:
            continue
            
        # Procesar metadatos (primeras 14 filas)
        if row[0] in metadata_keys:
            metadata[row[0]] = row[1] if len(row) > 1 else ""
            if row[0] == "Publish Time":
                publish_time = row[1] if len(row) > 1 else ""
        
        # Procesar encabezados de comentarios
        elif row[0] == 'Comment Number (ID)':
            continue  # Saltamos esta fila, usaremos nuestros propios headers
            
        # Procesar datos de comentarios
        elif len(row) > 0 and row[0].isdigit():
            # Verificar si necesitamos obtener seguidores
            if len(row) > 2:
                username = row[2]  # User @
                
                # Si no hay datos de seguidores o son N/A, obtenerlos
                followers_col = 8  # Posición de Followers en la estructura original
                if len(row) <= followers_col or row[followers_col] in ['', 'N/A'] or not str(row[followers_col]).isdigit():
                    followers = get_followers(username)
                    # Asegurar que la fila tenga suficientes columnas
                    while len(row) <= followers_col:
                        row.append('')
                    row[followers_col] = str(followers)
                    time.sleep(0.5)  # Pausa para evitar rate limiting
            
            comments_data.append(row)

# Crear el XLSX con la nueva estructura
print("Creando estructura del XLSX...")

# PASO 1: Crear headers horizontales para metadatos
metadata_headers = [
    "Now", "Post URL", "Publisher Nickname", "Publisher @", "Publisher URL",
    "Publish Time", "Post Likes", "Post Shares", "Description",
    "Number of 1st level comments", "Number of 2nd level comments",
    "Total Comments (actual)", "Total Comments (TikTok says)", "Difference"
]

# Agregar headers de metadatos en la primera fila
for col_num, header in enumerate(metadata_headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Agregar valores de metadatos en la segunda fila
for col_num, header in enumerate(metadata_headers, 1):
    original_key = [k for k in metadata_keys if k == header or k.startswith(header.split(' (')[0])]
    if original_key:
        value = metadata.get(original_key[0], "")
        ws.cell(row=2, column=col_num).value = value

# PASO 2: Dejar dos columnas vacías (columnas 15 y 16)

# PASO 3: Agregar headers de comentarios empezando en la columna 17
comment_headers = [
    "Comment Number (ID)", "Nickname", "User @", "User URL", "Comment Text",
    "Time", "Likes", "Profile Picture URL", "Followers", "Is 2nd Level Comment",
    "User Replied To", "Number of Replies"
]

start_col = 17  # Columna donde empiezan los datos de comentarios
for col_num, header in enumerate(comment_headers):
    cell = ws.cell(row=1, column=start_col + col_num)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# PASO 4: Agregar datos de comentarios
print(f"Agregando {len(comments_data)} comentarios...")
for row_num, comment_row in enumerate(comments_data, 3):  # Empezar en fila 3
    for col_num, value in enumerate(comment_row[:len(comment_headers)]):
        ws.cell(row=row_num, column=start_col + col_num).value = value

# Ajustar el ancho de las columnas
print("Ajustando ancho de columnas...")
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max(max_length + 2, 10), 50)  # Mínimo 10, máximo 50
    ws.column_dimensions[column_letter].width = adjusted_width

# Generar nombre de archivo con fecha del video
if publish_time:
    formatted_date = format_date_for_filename(publish_time)
    xlsx_filename = f"video_tiktok_{formatted_date}.xlsx"
else:
    xlsx_filename = f"video_tiktok_unknown_date.xlsx"

export_dir = path.join(cur_dir, "..", "exported data")
if not path.exists(export_dir):
    from os import makedirs
    makedirs(export_dir)
xlsx_path = path.join(export_dir, xlsx_filename)

# Guardar XLSX
print(f"Guardando archivo XLSX...")
wb.save(xlsx_path)
wb.close()

# Eliminar CSV original
try:
    remove(csv_path)
    print("Archivo CSV temporal eliminado.")
except:
    pass

print(f"\x1b[32m[*]\x1b[0m XLSX generado exitosamente: {xlsx_path}")
print(f"Total de comentarios procesados: {len(comments_data)}")
print("Estructura del archivo:")
print("- Fila 1: Headers de metadatos (columnas 1-14) + Headers de comentarios (columnas 17-28)")
print("- Fila 2: Valores de metadatos")
print("- Fila 3+: Datos de comentarios")
print("- Columnas 15-16: Vacías (separador)")